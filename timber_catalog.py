# Name: Laurence Finn
# Date: 2/25/2023
# Description: Module for the Timber cataloging classes and functions.

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm
from timber_settings import TimberSettings
from timber_logger import TimberLogger


class TimberCatalog:

    def __init__(self):
        self.directory = ""
        self.ignored_directories = []

        # initialize settings
        self.settings = TimberSettings()
        self.settings.load_settings()

        # initialize logger
        self.logger = TimberLogger()
        self.logger.log("Timber Catalog initialized.")

    def set_directory(self, directory="", command_line=False):
        if command_line:
            self.directory = directory
            if not os.path.isdir(self.directory):
                msg = "Invalid source directory. Exiting..."
                print(msg), self.logger.log(msg)
                sys.exit(1)
            return

        preloaded = False
        if len(self.settings.recent_catalogs) > 0:
            print("Recently cataloged directories:")
            for i in range(len(self.settings.recent_catalogs)):
                print(f"{i + 1}. {self.settings.recent_catalogs[i]}")
            while True:
                print("Enter a number to reuse one of the above, or press enter to use a different directory.")
                choice = int(input("> "))
                if choice == "":
                    break
                elif 0 < int(choice) <= len(self.settings.recent_catalogs):
                    self.directory = self.settings.recent_catalogs[choice - 1]
                    preloaded = True
                    break
                else:
                    print("Invalid choice. Please try again.")
                    continue

        if not preloaded:
            # Get the directory to list
            while True:
                self.directory = input("Enter the directory to catalog. \n"
                                       "Example: C:\\Users\\Laurence\\Desktop\\Media\n"
                                       "> ")
                if os.path.isdir(self.directory):
                    break
                else:
                    print("That is not a valid directory. Please try again.")

        self.settings.add_catalog(self.directory)
        self.settings.save_settings()

    def set_ignored_directories(self, ignored_directories="", command_line=False):

        if not command_line:
            self.ignored_directories = input("Enter a list of directories to ignore (separated by commas),"
                                             "or press enter to include all directories. \n"
                                             "Directories starting with a period (.) are ignored by default.\n"
                                             "Do NOT use spaces.\n"
                                             "Example: AppData,Music,Videos\\Cats\n"
                                             "> ")
            self.ignored_directories = self.ignored_directories.split(",")
        else:
            self.ignored_directories = ignored_directories.split(",")

        self.ignored_directories.append("$RECYCLE.BIN")

        # Add directories starting with a period to the ignored directories list
        for root, dirs, files in os.walk(self.directory):
            for d in dirs:
                if d.startswith('.'):
                    self.ignored_directories.append(d)

    def create_workbook(self, command_line=False):
        # Create a new workbook and select active worksheet
        workbook = Workbook()
        worksheet = workbook.active

        # Set column headers
        worksheet.append(["Directory", "Filename", "Size", "Time"])

        # Set column headers to light blue background
        for cell in worksheet["1:1"]:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Set the headers to freeze
        worksheet.freeze_panes = 'A2'

        print("Cataloging files... (This may take a while.)")

        # count the number of files
        file_count = 0
        for root, dirs, files in os.walk(self.directory):
            dirs[:] = [d for d in dirs if d not in self.ignored_directories]
            for file in files:
                if not file.startswith('.'):
                    file_count += 1

        with tqdm(total=file_count, unit='file') as pbar:
            pbar.format_meter(n=0, total=file_count, elapsed=0, prefix='Progress', unit='file',
                              colour='#00ff00')

            for root, dirs, files in os.walk(self.directory):
                dirs[:] = [d for d in dirs if d not in self.ignored_directories]

                for filename in files:
                    if not filename.startswith('.'):
                        # Get the file's path, split into directory and file name, without the source directory
                        file_path = os.path.join(root, filename)
                        directory_name, file_name = os.path.split(file_path)
                        directory_name = directory_name.replace(self.directory, "")

                        # If the directory name is bytes type, decode it to utf-8
                        if isinstance(directory_name, bytes):
                            directory_name = directory_name.decode('utf-8')

                        # If the file name is bytes type, decode it to utf-8
                        if isinstance(file_name, bytes):
                            file_name = file_name.decode('utf-8')

                        # Get the file's size and convert it to MB or GB
                        file_size = os.path.getsize(file_path)
                        if file_size < 1_073_741_824:
                            file_size = f"{file_size / 1_048_576:.2f} MB"
                        else:
                            file_size = f"{file_size / 1_073_741_824:.2f} GB"
                        file_size = str(file_size.encode('utf-8'))
                        file_size = file_size.lstrip("b'")
                        file_size = file_size.replace("\\", "/")
                        file_size = file_size.replace("//", "/")
                        file_size = file_size.replace("b\"", "")
                        file_size = file_size.replace("\"", "")
                        file_size = file_size.rstrip("'")

                        # Get the file's modification time and format it
                        file_mtime = os.path.getmtime(file_path)
                        mtime_dt = datetime.date.fromtimestamp(file_mtime)
                        mtime_str = mtime_dt.strftime('%Y-%m-%d %H:%M:%S')
                        mtime_str = str(mtime_str.encode('utf-8'))
                        mtime_str = mtime_str.replace("b'", "")
                        mtime_str = mtime_str.replace("'", "")

                        # Write the file's information to the worksheet
                        worksheet.append([directory_name, file_name, file_size, mtime_str])
                        pbar.update(1)

        print("Formatting the catalog file...")

        # for each row, set the odd rows to light gray and the even rows to white
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if (cell.row % 2 == 1) and (cell.row != 1):
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                elif (cell.row % 2 == 0) and (cell.row != 1):
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Set column widths
        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["B"].width = 40
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 20

        # Create a file name called "Catalog of" and the base directory name.

        # If the directory name is just a drive letter, use the drive letter as the workbook_filename
        if self.directory.endswith(":") or self.directory.endswith(":\\"):
            workbook_filename = self.directory.rstrip(":\\")
        else:
            workbook_filename = os.path.basename(self.directory.rstrip("\\/"))
        workbook_filename = "Catalog of " + workbook_filename + ".xlsx"

        # if the file exists, prompt the user for overwrite confirmation
        # if command line, skip this
        if os.path.isfile(workbook_filename) and not command_line:
            overwrite = input("A file named \"" + workbook_filename + "\" already exists. Overwrite? (y/n): ")
            if overwrite.lower() != "y":
                while True:
                    print("Enter a new file name: ")
                    workbook_filename = input()
                    if not all(c in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_-. "
                               for c in workbook_filename):
                        print("The filename contains illegal characters. Please try again.\n"
                              "A filename can't contain any of the following characters:\n"
                              " / \\ : * ? \" < > |")
                        continue
                    else:
                        workbook_filename += ".xlsx"
                        break

        # Try to save the workbook, and if it fails, print an error message
        try:
            workbook.save(workbook_filename)
            print("\"" + workbook_filename + "\" created successfully.")
            self.logger.log("Catalog created successfully.")
        except PermissionError:
            msg = "Error: Could not write to file. File is either open in another program or is read-only. \n" \
                    "Please close " + workbook_filename + " before running this program again."
            print(msg), self.logger.log(msg)
            exit(1)

    def disclaimer(self):
        print("===============================================================")
        print("================Timber Tree Cataloging Tool====================")
        print("===============================================================")
        print("This program will create a catalog of all files in a directory.\n"
              "The catalog will be saved as an Excel file in the same location\n"
              "as this program.")
        print("===============================================================")

    def catalog(self, directory="", ignored_directories="", command_line=False):
        if command_line:
            self.set_directory(directory, True)
            self.set_ignored_directories(ignored_directories, True)
            self.create_workbook(True)
        else:
            self.disclaimer()
            self.set_directory()
            self.set_ignored_directories()
            self.create_workbook()
